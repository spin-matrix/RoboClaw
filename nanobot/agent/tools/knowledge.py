"""SQLite knowledge base tools for structured file import and query."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from nanobot.agent.tools.base import Tool
from nanobot.utils.helpers import ensure_dir, safe_filename


BLOCKED_SQL = {
    "insert",
    "update",
    "delete",
    "drop",
    "alter",
    "create",
    "replace",
    "attach",
    "detach",
    "pragma",
    "vacuum",
}


def knowledge_root(workspace: Path) -> Path:
    return ensure_dir(workspace / "knowledge")


def knowledge_db_path(workspace: Path) -> Path:
    return knowledge_root(workspace) / "db.sqlite"


def knowledge_uploads_dir(workspace: Path) -> Path:
    return ensure_dir(knowledge_root(workspace) / "uploads")


def _slug(text: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_]+", "_", text.strip()).strip("_").lower()
    return text or "col"


def _unique_names(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in names:
        name = _slug(raw)
        idx = seen.get(name, 0)
        seen[name] = idx + 1
        out.append(name if idx == 0 else f"{name}_{idx}")
    return out


def parse_date(value: str) -> str | None:
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_number(value: str) -> int | float | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    cleaned = (
        cleaned.replace(",", "")
        .replace("，", "")
        .replace("元", "")
        .replace("伏", "")
        .replace("赫兹", "")
        .replace("升", "")
        .replace("千克", "")
        .replace("分贝", "")
        .replace("毫米", "")
        .replace("每分钟", "")
        .replace("转", "")
        .replace("千瓦时", "")
    )
    if re.fullmatch(r"-?\d+", cleaned):
        return int(cleaned)
    if re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
        return float(cleaned)
    return None


def parse_bool(value: str) -> int | None:
    cleaned = value.strip().lower()
    mapping = {
        "是": 1,
        "否": 0,
        "true": 1,
        "false": 0,
        "yes": 1,
        "no": 0,
        "1": 1,
        "0": 0,
    }
    return mapping.get(cleaned)


def load_rows(path: Path, sheet: str | None = None) -> tuple[list[dict[str, Any]], str | None]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, list) or not all(isinstance(row, dict) for row in data):
            raise ValueError("JSON input must be an array of objects")
        return data, None
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f)), None
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for xlsx import") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        matrix = list(ws.iter_rows(values_only=True))
        if not matrix:
            return [], ws.title
        headers = [str(v).strip() if v is not None else "" for v in matrix[0]]
        rows: list[dict[str, Any]] = []
        for row in matrix[1:]:
            item = {}
            for idx, value in enumerate(row):
                key = headers[idx] if idx < len(headers) and headers[idx] else f"column_{idx+1}"
                item[key] = value
            rows.append(item)
        return rows, ws.title
    raise ValueError(f"Unsupported file type: {suffix}")


def create_meta_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS datasets (
            dataset_id TEXT PRIMARY KEY,
            table_name TEXT NOT NULL,
            source_file TEXT NOT NULL,
            file_type TEXT NOT NULL,
            sheet_name TEXT,
            row_count INTEGER NOT NULL,
            session_key TEXT,
            imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS columns_registry (
            table_name TEXT NOT NULL,
            column_name TEXT NOT NULL,
            source_column TEXT NOT NULL,
            storage_kind TEXT NOT NULL,
            sqlite_type TEXT NOT NULL,
            semantic_type TEXT NOT NULL,
            PRIMARY KEY (table_name, column_name)
        );
        """
    )


def build_schema(rows: list[dict[str, Any]]) -> tuple[list[tuple[str, str, str, str]], list[str]]:
    source_columns: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in source_columns:
                source_columns.append(key)
    sanitized = _unique_names(source_columns)
    schema: list[tuple[str, str, str, str]] = []
    insert_columns: list[str] = []
    for source_name, col in zip(source_columns, sanitized):
        schema.append((col, source_name, "raw", "TEXT"))
        insert_columns.append(col)
        non_empty: list[str] = []
        for row in rows[:50]:
            value = row.get(source_name)
            if value is None:
                continue
            text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False)
            if str(text).strip():
                non_empty.append(str(text))
        if non_empty and all(parse_bool(v) is not None for v in non_empty):
            schema.append((f"{col}__norm", source_name, "normalized", "INTEGER"))
            insert_columns.append(f"{col}__norm")
        elif non_empty and all(parse_number(v) is not None for v in non_empty):
            schema.append((f"{col}__norm", source_name, "normalized", "REAL"))
            insert_columns.append(f"{col}__norm")
        elif non_empty and all(parse_date(v) is not None for v in non_empty):
            schema.append((f"{col}__date", source_name, "normalized", "TEXT"))
            insert_columns.append(f"{col}__date")
    return schema, insert_columns


def row_values(row: dict[str, Any], schema: list[tuple[str, str, str, str]]) -> list[object]:
    values: list[object] = []
    for col_name, source_name, storage_kind, _ in schema:
        raw = row.get(source_name)
        if storage_kind == "raw":
            if raw is None:
                values.append(None)
            elif isinstance(raw, (dict, list)):
                values.append(json.dumps(raw, ensure_ascii=False))
            else:
                values.append(str(raw))
            continue
        raw_text = "" if raw is None else (raw if isinstance(raw, str) else json.dumps(raw, ensure_ascii=False))
        if col_name.endswith("__date"):
            values.append(parse_date(raw_text))
        elif col_name.endswith("__norm"):
            parsed_bool = parse_bool(raw_text)
            values.append(parsed_bool if parsed_bool is not None else parse_number(raw_text))
        else:
            values.append(None)
    return values


def _semantic_type(col_name: str, sqlite_type: str) -> str:
    if col_name.endswith("__date"):
        return "date"
    if col_name.endswith("__norm") and sqlite_type == "REAL":
        return "number"
    if col_name.endswith("__norm") and sqlite_type == "INTEGER":
        return "boolean"
    return "text"


def create_table(conn: sqlite3.Connection, table_name: str, schema: list[tuple[str, str, str, str]]) -> None:
    cols = ", ".join(f'"{name}" {sql_type}' for name, _, _, sql_type in schema)
    conn.execute(f'CREATE TABLE "{table_name}" ({cols})')


def build_names(input_path: Path, table_name: str | None = None) -> tuple[str, str, str]:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    digest = hashlib.sha1(f"{input_path.name}:{datetime.now(timezone.utc).isoformat()}".encode("utf-8")).hexdigest()[:8]
    return f"ds_{stamp}_{digest}", table_name or f"kb_{stamp}_{digest}", datetime.now(timezone.utc).isoformat()


def connect_db(db_path: Path) -> sqlite3.Connection:
    ensure_dir(db_path.parent)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def import_file_to_db(
    db_path: Path,
    input_path: Path,
    sheet: str | None = None,
    session_key: str | None = None,
    table_name: str | None = None,
) -> dict[str, Any]:
    rows, sheet_name = load_rows(input_path, sheet)
    if not rows:
        raise ValueError("Input file contains no data rows")
    dataset_id, actual_table, imported_at = build_names(input_path, table_name)
    conn = connect_db(db_path)
    try:
        create_meta_tables(conn)
        schema, insert_columns = build_schema(rows)
        create_table(conn, actual_table, schema)
        placeholders = ", ".join("?" for _ in insert_columns)
        quoted_columns = ", ".join(f'"{c}"' for c in insert_columns)
        insert_sql = f'INSERT INTO "{actual_table}" ({quoted_columns}) VALUES ({placeholders})'
        conn.executemany(insert_sql, [row_values(row, schema) for row in rows])
        conn.execute(
            """
            INSERT INTO datasets (
                dataset_id, table_name, source_file, file_type, sheet_name, row_count, session_key, imported_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dataset_id,
                actual_table,
                input_path.name,
                input_path.suffix.lower().lstrip("."),
                sheet_name,
                len(rows),
                session_key,
                imported_at,
            ),
        )
        conn.executemany(
            """
            INSERT INTO columns_registry (
                table_name, column_name, source_column, storage_kind, sqlite_type, semantic_type
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                (actual_table, col_name, source_name, storage_kind, sqlite_type, _semantic_type(col_name, sqlite_type))
                for col_name, source_name, storage_kind, sqlite_type in schema
            ],
        )
        conn.commit()
    finally:
        conn.close()
    return {
        "dataset_id": dataset_id,
        "table_name": actual_table,
        "source_file": input_path.name,
        "row_count": len(rows),
        "sheet_name": sheet_name,
        "db_path": str(db_path),
    }


def save_uploaded_knowledge_file(workspace: Path, file_name: str, raw: bytes) -> Path:
    uploads_dir = knowledge_uploads_dir(workspace)
    digest = hashlib.sha1(raw).hexdigest()[:12]
    safe_name = safe_filename(Path(file_name).stem)
    suffix = Path(file_name).suffix or ".bin"
    stored = uploads_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{digest}_{safe_name}{suffix}"
    stored.write_bytes(raw)
    return stored


def list_datasets(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        "SELECT dataset_id, table_name, source_file, file_type, row_count, session_key, imported_at "
        "FROM datasets ORDER BY imported_at DESC"
    ).fetchall()
    return [dict(row) for row in rows]


def dataset_display_name(source_file: str, table_name: str) -> str:
    stem = Path(source_file).stem.strip()
    stem = re.sub(r"^\d{8}_\d{6}_[0-9a-f]{12}_", "", stem)
    return stem or table_name


def summarize_dataset(conn: sqlite3.Connection, table_name: str) -> str:
    cols = conn.execute(
        "SELECT source_column, semantic_type, storage_kind FROM columns_registry WHERE table_name = ? ORDER BY rowid",
        (table_name,),
    ).fetchall()
    source_cols = [row["source_column"] for row in cols if row["storage_kind"] == "raw"]
    semantic_cols = [row["source_column"] for row in cols if row["semantic_type"] in {"number", "date", "boolean"}]
    preview = "、".join(source_cols[:4]) if source_cols else "字段未知"
    if semantic_cols:
        return f"存储 {preview} 等字段，可用于结构化查询。"
    return f"存储 {preview} 等字段。"


def list_datasets_indexed(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    datasets = list_datasets(conn)
    indexed: list[dict[str, Any]] = []
    for idx, item in enumerate(datasets, start=1):
        row = dict(item)
        row["index"] = idx
        row["display_name"] = dataset_display_name(item["source_file"], item["table_name"])
        row["summary"] = summarize_dataset(conn, item["table_name"])
        indexed.append(row)
    return indexed


def render_datasets_catalog(conn: sqlite3.Connection) -> str:
    datasets = list_datasets_indexed(conn)
    if not datasets:
        return "当前知识库中没有已导入的数据表。"
    lines = [f"当前知识库共有 {len(datasets)} 张表：", "", "索引 | 表名 | 描述"]
    for row in datasets:
        lines.append(f"{row['index']} | {row['display_name']} | {row['summary']}")
    return "\n".join(lines)


def describe_table(conn: sqlite3.Connection, table: str) -> dict[str, Any]:
    dataset = conn.execute(
        "SELECT dataset_id, table_name, source_file, file_type, row_count, session_key, imported_at, sheet_name "
        "FROM datasets WHERE table_name = ? ORDER BY imported_at DESC LIMIT 1",
        (table,),
    ).fetchone()
    columns = conn.execute(
        "SELECT column_name, source_column, storage_kind, sqlite_type, semantic_type "
        "FROM columns_registry WHERE table_name = ? ORDER BY column_name",
        (table,),
    ).fetchall()
    return {"dataset": dict(dataset) if dataset else None, "columns": [dict(row) for row in columns]}


def schema_context(conn: sqlite3.Connection, table: str) -> str:
    info = describe_table(conn, table)
    if not info["dataset"]:
        raise ValueError(f"Unknown table: {table}")
    ds = info["dataset"]
    lines = [
        f"Table: {ds['table_name']}",
        f"Source file: {ds['source_file']}",
        f"File type: {ds['file_type']}",
        f"Row count: {ds['row_count']}",
        "Columns:",
    ]
    for col in info["columns"]:
        lines.append(
            f"- {col['column_name']} ({col['sqlite_type']}, {col['semantic_type']}, {col['storage_kind']}; source={col['source_column']})"
        )
    return "\n".join(lines)


def render_dataset_detail(conn: sqlite3.Connection, table: str) -> str:
    info = describe_table(conn, table)
    if not info["dataset"]:
        raise ValueError(f"Unknown table: {table}")
    ds = info["dataset"]
    raw_columns = [col["source_column"] for col in info["columns"] if col["storage_kind"] == "raw"]
    structured_columns = [
        col["source_column"] for col in info["columns"] if col["semantic_type"] in {"number", "date", "boolean"}
    ]
    display_name = dataset_display_name(ds["source_file"], ds["table_name"])
    lines = [
        f"表名: {display_name}",
        f"数据量: {ds['row_count']} 行",
        f"来源文件: {ds['source_file']}",
        f"描述: {summarize_dataset(conn, table)}",
        "主要字段:",
    ]
    for col in raw_columns[:12]:
        lines.append(f"- {col}")
    if structured_columns:
        lines.append("可用于筛选/排序的字段:")
        for col in structured_columns[:8]:
            lines.append(f"- {col}")
    lines.append(f"内部表名: {ds['table_name']}")
    return "\n".join(lines)


def sample_rows_context(conn: sqlite3.Connection, table: str, limit: int = 3) -> str:
    info = describe_table(conn, table)
    if not info["dataset"]:
        raise ValueError(f"Unknown table: {table}")
    raw_columns = [col["column_name"] for col in info["columns"] if col["storage_kind"] == "raw"]
    if not raw_columns:
        return f"Table: {table}\nSample rows: none"
    select_cols = ", ".join(f'"{col}"' for col in raw_columns[:8])
    rows = conn.execute(f'SELECT {select_cols} FROM "{table}" LIMIT ?', (limit,)).fetchall()
    payload = [dict(row) for row in rows]
    return f"Table: {table}\nSample rows:\n{json.dumps(payload, ensure_ascii=False, indent=2)}"


def ensure_safe_select(sql: str) -> str:
    text = sql.strip().rstrip(";")
    if ";" in text:
        raise ValueError("Only one SQL statement is allowed")
    if not re.match(r"^\s*select\b", text, flags=re.I):
        raise ValueError("Only SELECT statements are allowed")
    lowered = text.lower()
    for token in BLOCKED_SQL:
        if re.search(rf"\b{re.escape(token)}\b", lowered):
            raise ValueError(f"Blocked SQL token: {token}")
    return text


def run_query(conn: sqlite3.Connection, sql: str, limit: int = 50) -> dict[str, Any]:
    safe_sql = ensure_safe_select(sql)
    wrapped = f"SELECT * FROM ({safe_sql}) LIMIT {int(limit)}"
    rows = conn.execute(wrapped).fetchall()
    return {"row_count": len(rows), "rows": [dict(row) for row in rows]}


def delete_dataset(conn: sqlite3.Connection, table_name: str) -> dict[str, Any]:
    dataset = conn.execute(
        "SELECT dataset_id, table_name, source_file FROM datasets WHERE table_name = ? ORDER BY imported_at DESC LIMIT 1",
        (table_name,),
    ).fetchone()
    if not dataset:
        raise ValueError(f"Unknown table: {table_name}")
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    conn.execute("DELETE FROM columns_registry WHERE table_name = ?", (table_name,))
    conn.execute("DELETE FROM datasets WHERE table_name = ?", (table_name,))
    conn.commit()
    return {
        "dataset_id": dataset["dataset_id"],
        "table_name": dataset["table_name"],
        "source_file": dataset["source_file"],
        "deleted": True,
    }


def delete_dataset_by_index(conn: sqlite3.Connection, index: int) -> dict[str, Any]:
    datasets = list_datasets_indexed(conn)
    for row in datasets:
        if row["index"] == index:
            return delete_dataset(conn, row["table_name"])
    raise ValueError(f"Unknown dataset index: {index}")


class _KnowledgeTool(Tool):
    def __init__(self, workspace: Path):
        self.workspace = workspace
        self.db_path = knowledge_db_path(workspace)


class KnowledgeImportTool(_KnowledgeTool):
    @property
    def name(self) -> str:
        return "knowledge_import"

    @property
    def description(self) -> str:
        return "Import a structured CSV, JSON, or XLSX file from the workspace into the SQLite knowledge base."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Path to csv/json/xlsx file, absolute or relative to workspace"},
                "sheet": {"type": "string", "description": "Optional sheet name for xlsx"},
                "table_name": {"type": "string", "description": "Optional explicit table name"},
                "session_key": {"type": "string", "description": "Optional session key for dataset metadata"},
            },
            "required": ["path"],
        }

    async def execute(
        self,
        path: str,
        sheet: str | None = None,
        table_name: str | None = None,
        session_key: str | None = None,
        **kwargs: Any,
    ) -> str:
        input_path = Path(path).expanduser()
        if not input_path.is_absolute():
            input_path = self.workspace / input_path
        if not input_path.exists():
            return f"Error: File not found: {input_path}"
        result = import_file_to_db(
            db_path=self.db_path,
            input_path=input_path.resolve(),
            sheet=sheet,
            session_key=session_key,
            table_name=table_name,
        )
        return json.dumps(result, ensure_ascii=False, indent=2)


class KnowledgeSchemaTool(_KnowledgeTool):
    @property
    def name(self) -> str:
        return "knowledge_schema"

    @property
    def description(self) -> str:
        return "List imported datasets or return model-facing schema context for a specific table."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["list", "describe", "schema_context", "catalog", "detail"],
                    "description": "Choose list/catalog for user-facing inventory, detail for one table summary, describe for JSON metadata, or schema_context for prompt-facing text.",
                },
                "table": {"type": "string", "description": "Required for describe or schema_context"},
            },
            "required": ["action"],
        }

    async def execute(self, action: str, table: str | None = None, **kwargs: Any) -> str:
        if not self.db_path.exists():
            return f"Error: Knowledge DB not found: {self.db_path}"
        conn = connect_db(self.db_path)
        try:
            if action == "list":
                return json.dumps(list_datasets_indexed(conn), ensure_ascii=False, indent=2)
            if action == "catalog":
                return render_datasets_catalog(conn)
            if not table:
                return "Error: table is required for this action"
            if action == "detail":
                return render_dataset_detail(conn, table)
            if action == "describe":
                return json.dumps(describe_table(conn, table), ensure_ascii=False, indent=2)
            return schema_context(conn, table)
        finally:
            conn.close()


class KnowledgeQueryTool(_KnowledgeTool):
    @property
    def name(self) -> str:
        return "knowledge_query"

    @property
    def description(self) -> str:
        return "Run a safe read-only SQLite SELECT query against the structured knowledge base."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "sql": {"type": "string", "description": "One SQLite SELECT statement"},
                "limit": {"type": "integer", "description": "Maximum rows to return", "minimum": 1, "maximum": 200},
            },
            "required": ["sql"],
        }

    async def execute(self, sql: str, limit: int = 50, **kwargs: Any) -> str:
        if not self.db_path.exists():
            return f"Error: Knowledge DB not found: {self.db_path}"
        conn = connect_db(self.db_path)
        try:
            return json.dumps(run_query(conn, sql, limit), ensure_ascii=False, indent=2)
        finally:
            conn.close()


class KnowledgeRemoveTool(_KnowledgeTool):
    @property
    def name(self) -> str:
        return "knowledge_remove"

    @property
    def description(self) -> str:
        return "Remove one imported dataset by table name from the SQLite knowledge base."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "table": {"type": "string", "description": "The imported table name to delete"},
                "index": {"type": "integer", "description": "The displayed dataset index to delete", "minimum": 1},
            },
            "required": [],
        }

    async def execute(self, table: str | None = None, index: int | None = None, **kwargs: Any) -> str:
        if not self.db_path.exists():
            return f"Error: Knowledge DB not found: {self.db_path}"
        if not table and index is None:
            return "Error: either table or index is required"
        conn = connect_db(self.db_path)
        try:
            if index is not None:
                return json.dumps(delete_dataset_by_index(conn, index), ensure_ascii=False, indent=2)
            return json.dumps(delete_dataset(conn, table or ""), ensure_ascii=False, indent=2)
        finally:
            conn.close()
